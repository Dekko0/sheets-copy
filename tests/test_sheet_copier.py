"""Testes do núcleo de cópia (core.sheet_copier)."""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName

from config.defaults import COPY_MODE_OVERLAY, COPY_MODE_REPLACE
from core.exceptions import SheetNotFoundError
from core.sheet_copier import copy_defined_names, copy_sheets
from tests.conftest import roundtrip


# --------------------------------------------------------------------------- #
# Conteúdo: valores, fórmulas e estilos
# --------------------------------------------------------------------------- #
def test_copies_values_and_formulas(loaded_template, destination):
    copy_sheets(loaded_template, destination, ["Dados"])
    ws = destination["Dados"]
    assert ws["A1"].value == "Item"
    assert ws["B2"].value == 100
    assert ws["B4"].value == "=SUM(B2:B3)"  # fórmula preservada como string


def test_copies_styles_and_number_format(loaded_template, destination):
    copy_sheets(loaded_template, destination, ["Dados"])
    ws = destination["Dados"]
    assert ws["A1"].font.bold is True
    assert ws["B1"].font.bold is True
    assert ws["A1"].alignment.horizontal == "center"
    assert ws["A1"].fill.fill_type == "solid"
    assert "2E7D32" in (ws["A1"].fill.fgColor.rgb or "")
    assert ws["B4"].number_format == "#,##0.00"


def test_copies_comment_and_hyperlink(loaded_template, destination):
    copy_sheets(loaded_template, destination, ["Dados"])
    ws = destination["Dados"]
    assert ws["A1"].comment is not None
    assert ws["A1"].comment.text == "Cabeçalho"
    assert ws["D1"].hyperlink is not None
    assert ws["D1"].hyperlink.target == "https://example.com"


# --------------------------------------------------------------------------- #
# Estruturas: mescladas, validações, formatação condicional, dimensões
# --------------------------------------------------------------------------- #
def test_copies_merged_cells(loaded_template, destination):
    copy_sheets(loaded_template, destination, ["Dados"])
    ws = destination["Dados"]
    assert "A6:C6" in [str(r) for r in ws.merged_cells.ranges]


def test_copies_data_validation(loaded_template, destination):
    copy_sheets(loaded_template, destination, ["Dados"])
    dvs = destination["Dados"].data_validations.dataValidation
    assert len(dvs) == 1
    assert dvs[0].type == "list"
    assert dvs[0].formula1 == '"Sim,Não"'
    assert "C2:C10" in str(dvs[0].sqref)


def test_copies_conditional_formatting(loaded_template, destination):
    copy_sheets(loaded_template, destination, ["Dados"])
    cf_rules = destination["Dados"].conditional_formatting._cf_rules
    assert sum(len(v) for v in cf_rules.values()) == 1


def test_copies_dimensions(loaded_template, destination):
    copy_sheets(loaded_template, destination, ["Dados"])
    ws = destination["Dados"]
    assert ws.column_dimensions["A"].width == 25
    assert ws.row_dimensions[1].height == 30


def test_copies_layout(loaded_template, destination):
    copy_sheets(loaded_template, destination, ["Dados"])
    ws = destination["Dados"]
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref == "A1:B4"
    assert ws.sheet_properties.tabColor is not None
    assert ws.sheet_properties.tabColor.rgb.endswith("FF0000")


def test_copies_chart(loaded_template, destination):
    # openpyxl 3.1.5 lê gráficos do arquivo, então loaded_template tem 1 gráfico.
    result = copy_sheets(loaded_template, destination, ["Dados"]).results[0]
    assert len(destination["Dados"]._charts) == 1
    assert result.charts == 1


# --------------------------------------------------------------------------- #
# Modos de cópia
# --------------------------------------------------------------------------- #
def test_replace_mode_wipes_extra_cells(loaded_template, destination):
    assert destination["Dados"]["A20"].value == "EXTRA_DESTINO"
    copy_sheets(loaded_template, destination, ["Dados"], mode=COPY_MODE_REPLACE)
    assert destination["Dados"]["A20"].value is None  # aba recriada do zero
    assert destination["Dados"]["A1"].value == "Item"


def test_overlay_mode_keeps_extra_cells(loaded_template, destination):
    copy_sheets(loaded_template, destination, ["Dados"], mode=COPY_MODE_OVERLAY)
    assert destination["Dados"]["A20"].value == "EXTRA_DESTINO"  # preservada
    assert destination["Dados"]["A1"].value == "Item"  # sobrescrita


def test_replace_preserves_sheet_position(loaded_template, destination):
    copy_sheets(loaded_template, destination, ["Dados"], mode=COPY_MODE_REPLACE)
    assert destination.sheetnames.index("Dados") == 0
    assert "Outra" in destination.sheetnames


def test_replace_when_target_has_single_sheet(loaded_template):
    dest = Workbook()
    dest.active.title = "Dados"
    copy_sheets(loaded_template, dest, ["Dados"], mode=COPY_MODE_REPLACE)
    assert dest.sheetnames == ["Dados"]
    assert dest["Dados"]["A1"].value == "Item"


# --------------------------------------------------------------------------- #
# Criação de aba inexistente e relatório
# --------------------------------------------------------------------------- #
def test_creates_missing_sheet(loaded_template, destination):
    assert "Calc" not in destination.sheetnames
    report = copy_sheets(loaded_template, destination, ["Calc"])
    assert "Calc" in destination.sheetnames
    assert report.results[0].created is True


def test_report_counts(loaded_template, destination):
    report = copy_sheets(loaded_template, destination, ["Dados", "Calc"])
    assert report.sheets_copied == 2
    assert report.sheets_created == 1  # só 'Calc' não existia


def test_progress_callback_is_called(loaded_template, destination):
    calls: list[tuple[int, int, str]] = []
    copy_sheets(
        loaded_template,
        destination,
        ["Dados", "Calc"],
        progress_callback=lambda cur, total, name: calls.append((cur, total, name)),
    )
    assert calls == [(1, 2, "Dados"), (2, 2, "Calc")]


# --------------------------------------------------------------------------- #
# Named ranges
# --------------------------------------------------------------------------- #
def test_copies_defined_names(loaded_template, destination):
    report = copy_sheets(loaded_template, destination, ["Dados"])
    assert "ListaSimNao" in destination.defined_names
    assert "ListaSimNao" in report.added_named_ranges


def test_defined_names_not_overwritten(loaded_template, destination):
    destination.defined_names.add(DefinedName("ListaSimNao", attr_text="Dados!$Z$99"))
    added = copy_defined_names(loaded_template, destination)
    assert "ListaSimNao" not in added
    assert destination.defined_names["ListaSimNao"].value == "Dados!$Z$99"


# --------------------------------------------------------------------------- #
# Erros e validação de entrada
# --------------------------------------------------------------------------- #
def test_raises_when_sheet_missing_in_template(loaded_template, destination):
    import pytest

    with pytest.raises(SheetNotFoundError):
        copy_sheets(loaded_template, destination, ["NaoExiste"])


def test_raises_on_invalid_mode(loaded_template, destination):
    import pytest

    with pytest.raises(ValueError):
        copy_sheets(loaded_template, destination, ["Dados"], mode="modo_invalido")


# --------------------------------------------------------------------------- #
# Round-trip de ponta a ponta
# --------------------------------------------------------------------------- #
def test_roundtrip_preserves_content(loaded_template, destination):
    copy_sheets(loaded_template, destination, ["Dados", "Calc"])
    reloaded = roundtrip(destination)
    ws = reloaded["Dados"]
    assert ws["B4"].value == "=SUM(B2:B3)"
    assert len(ws.data_validations.dataValidation) == 1
    assert "A6:C6" in [str(r) for r in ws.merged_cells.ranges]
    assert ws.column_dimensions["A"].width == 25
    assert "ListaSimNao" in reloaded.defined_names
    assert len(ws._charts) == 1  # gráfico sobrevive cópia + save + reload
