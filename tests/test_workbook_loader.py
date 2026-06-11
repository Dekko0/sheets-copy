"""Testes do carregamento/serialização de workbooks (core.workbook_loader)."""

from __future__ import annotations

import io

import pytest
from openpyxl import load_workbook

from core.exceptions import InvalidWorkbookError
from core.workbook_loader import (
    _should_keep_vba,
    load_workbook_from_bytes,
    load_workbook_from_path,
    save_workbook_to_bytes,
)
from tests.conftest import make_template


def test_load_from_path(sample_template_file):
    wb = load_workbook_from_path(sample_template_file)
    assert "Dados" in wb.sheetnames
    assert "Calc" in wb.sheetnames


def test_load_from_path_missing_file():
    with pytest.raises(InvalidWorkbookError):
        load_workbook_from_path("c:/caminho/que/nao/existe.xlsx")


def test_load_from_path_invalid_extension(tmp_path):
    bad = tmp_path / "arquivo.txt"
    bad.write_text("conteúdo qualquer", encoding="utf-8")
    with pytest.raises(InvalidWorkbookError):
        load_workbook_from_path(bad)


def test_load_from_bytes():
    buffer = io.BytesIO()
    make_template().save(buffer)
    wb = load_workbook_from_bytes(buffer.getvalue(), filename="modelo.xlsx")
    assert "Dados" in wb.sheetnames


def test_load_from_bytes_corrupt_content():
    with pytest.raises(InvalidWorkbookError):
        load_workbook_from_bytes(b"isto nao e um excel", filename="x.xlsx")


def test_load_from_bytes_invalid_extension():
    with pytest.raises(InvalidWorkbookError):
        load_workbook_from_bytes(b"qualquer", filename="x.csv")


def test_should_keep_vba_by_extension():
    assert _should_keep_vba("planilha.xlsm") is True
    assert _should_keep_vba("planilha.xlsx") is False
    assert _should_keep_vba("PLANILHA.XLSM") is True  # case-insensitive


def test_save_workbook_to_bytes_roundtrip():
    buffer = save_workbook_to_bytes(make_template())
    assert buffer.getbuffer().nbytes > 0
    wb = load_workbook(buffer)
    assert "Dados" in wb.sheetnames
