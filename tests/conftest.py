"""Fixtures e construtores de workbooks de teste.

Os workbooks são gerados programaticamente (sem depender do ``RD_MODELO.xlsx``
real), exercitando todos os recursos que o núcleo precisa copiar: fórmulas,
estilos, mescladas, validações, formatação condicional, comentário, hyperlink,
dimensões, gráfico, named range e dependências entre abas.
"""

from __future__ import annotations

import io
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

FIXTURES_DIR = Path(__file__).parent / "fixtures"


def roundtrip(wb: Workbook) -> Workbook:
    """Salva o workbook em memória e o recarrega (simula leitura de um arquivo)."""
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return load_workbook(buffer)


def make_template() -> Workbook:
    """Constrói um modelo rico em recursos, com duas abas: 'Dados' e 'Calc'."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"

    # Cabeçalho estilizado + comentário.
    ws["A1"] = "Item"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=12)
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor="2E7D32")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].comment = Comment("Cabeçalho", "Tester")
    ws["B1"] = "Valor"
    ws["B1"].font = Font(bold=True)

    # Dados + fórmula com formato numérico.
    ws["A2"], ws["B2"] = "Energia", 100
    ws["A3"], ws["B3"] = "Demanda", 200
    ws["A4"] = "Total"
    ws["B4"] = "=SUM(B2:B3)"
    ws["B4"].number_format = "#,##0.00"

    # Hyperlink.
    ws["D1"] = "site"
    ws["D1"].hyperlink = "https://example.com"

    # Mesclada.
    ws.merge_cells("A6:C6")
    ws["A6"] = "Resumo"

    # Dimensões, congelamento, cor da aba, filtro.
    ws.column_dimensions["A"].width = 25
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.sheet_properties.tabColor = "FF0000"
    ws.auto_filter.ref = "A1:B4"

    # Validação de dados (lista Sim/Não).
    dv = DataValidation(type="list", formula1='"Sim,Não"', allow_blank=True)
    dv.add("C2:C10")
    ws.add_data_validation(dv)

    # Formatação condicional.
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ws.conditional_formatting.add(
        "B2:B3", CellIsRule(operator="greaterThan", formula=["150"], fill=red)
    )

    # Gráfico.
    chart = BarChart()
    chart.title = "Consumo"
    chart.add_data(Reference(ws, min_col=2, min_row=2, max_row=3))
    ws.add_chart(chart, "E5")

    # Segunda aba com dependências cruzadas.
    calc = wb.create_sheet("Calc")
    calc["A1"] = "=Dados!B4 + FATURAS!A1"      # Dados (ok) + FATURAS (quebrada)
    calc["A2"] = "='Análise Econômica'!B2"     # quotada + acento (quebrada)
    calc["A3"] = '="texto com ! interno"'      # literal: NÃO é dependência
    calc["A4"] = "=Calc!A1"                     # auto-referência: ignorada
    calc["A5"] = "=FATURAS!Z9"                  # FATURAS de novo (dedupe)

    # Named range (workbook-level).
    wb.defined_names.add(DefinedName("ListaSimNao", attr_text="Dados!$C$1:$C$10"))
    return wb


def make_destination() -> Workbook:
    """Destino com 'Dados' (conteúdo antigo + célula extra) e 'Outra'; sem 'Calc'."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"
    ws["A1"] = "valor antigo"
    ws["A20"] = "EXTRA_DESTINO"  # fora do intervalo do modelo
    wb.create_sheet("Outra")
    return wb


@pytest.fixture
def template() -> Workbook:
    """Modelo recém-construído em memória (gráfico com referência viva)."""
    return make_template()


@pytest.fixture
def loaded_template() -> Workbook:
    """Modelo após save+reload — espelha um arquivo real lido do disco."""
    return roundtrip(make_template())


@pytest.fixture
def destination() -> Workbook:
    """Planilha destino recém-construída."""
    return make_destination()


@pytest.fixture(scope="session")
def sample_template_file() -> Path:
    """Gera (uma vez) tests/fixtures/sample_template.xlsx e retorna o caminho."""
    FIXTURES_DIR.mkdir(parents=True, exist_ok=True)
    path = FIXTURES_DIR / "sample_template.xlsx"
    make_template().save(path)
    return path
