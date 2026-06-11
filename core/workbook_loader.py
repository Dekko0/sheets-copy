"""Carregamento de workbooks Excel a partir de caminho, bytes ou upload.

Encapsula o ``openpyxl.load_workbook`` aplicando regras do domínio:

* valida a extensão (``.xlsx`` / ``.xlsm``);
* ativa ``keep_vba`` automaticamente para ``.xlsm`` (preserva macros);
* nunca usa ``data_only`` por padrão, para **não perder as fórmulas**;
* converte qualquer falha de leitura em :class:`InvalidWorkbookError`.

O original em disco nunca é modificado: ``openpyxl`` trabalha sobre uma cópia
em memória.
"""

from __future__ import annotations

import io
import logging
from pathlib import Path
from typing import IO

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from config.defaults import ALLOWED_EXTENSIONS
from core.exceptions import InvalidWorkbookError

logger = logging.getLogger(__name__)


def _should_keep_vba(filename: str) -> bool:
    """Decide se as macros VBA devem ser preservadas, com base na extensão."""
    return Path(filename).suffix.lower() == ".xlsm"


def _validate_extension(filename: str) -> None:
    """Valida a extensão do arquivo.

    Args:
        filename: Nome (ou caminho) do arquivo.

    Raises:
        InvalidWorkbookError: Se a extensão não estiver em
            :data:`config.defaults.ALLOWED_EXTENSIONS`.
    """
    suffix = Path(filename).suffix.lower()
    if suffix not in ALLOWED_EXTENSIONS:
        permitidas = ", ".join(ALLOWED_EXTENSIONS)
        raise InvalidWorkbookError(
            f"Extensão '{suffix or '(vazia)'}' não suportada. "
            f"Use um dos formatos: {permitidas}."
        )


def load_workbook_from_path(
    path: str | Path,
    *,
    data_only: bool = False,
    keep_vba: bool | None = None,
) -> Workbook:
    """Carrega um workbook a partir de um caminho em disco.

    Args:
        path: Caminho do arquivo ``.xlsx`` ou ``.xlsm``.
        data_only: Se ``True``, lê os valores calculados em vez das fórmulas.
            Mantenha ``False`` para preservar fórmulas (padrão).
        keep_vba: Força preservar (ou não) macros. Se ``None``, decide
            automaticamente pela extensão.

    Returns:
        O :class:`openpyxl.Workbook` carregado.

    Raises:
        InvalidWorkbookError: Se o arquivo não existir, tiver extensão inválida
            ou não puder ser lido como Excel.
    """
    path = Path(path)
    if not path.is_file():
        raise InvalidWorkbookError(f"Arquivo não encontrado: {path}")

    _validate_extension(path.name)
    if keep_vba is None:
        keep_vba = _should_keep_vba(path.name)

    logger.info("Carregando workbook de %s (keep_vba=%s)", path, keep_vba)
    try:
        return load_workbook(filename=str(path), data_only=data_only, keep_vba=keep_vba)
    except InvalidFileException as exc:
        raise InvalidWorkbookError(f"Arquivo Excel inválido: {path.name}") from exc
    except PermissionError as exc:
        raise InvalidWorkbookError(
            f"Sem permissão para ler o arquivo: {path.name}. "
            "Feche-o no Excel e tente novamente."
        ) from exc
    except Exception as exc:  # noqa: BLE001 - normaliza qualquer falha de leitura
        raise InvalidWorkbookError(
            f"Não foi possível abrir '{path.name}': {exc}"
        ) from exc


def load_workbook_from_bytes(
    data: bytes | IO[bytes],
    *,
    filename: str = "upload.xlsx",
    data_only: bool = False,
    keep_vba: bool | None = None,
) -> Workbook:
    """Carrega um workbook a partir de bytes (ex.: upload do Streamlit).

    Args:
        data: Conteúdo binário do arquivo, ou um objeto file-like de bytes.
        filename: Nome original, usado para validar a extensão e decidir
            ``keep_vba``.
        data_only: Veja :func:`load_workbook_from_path`.
        keep_vba: Veja :func:`load_workbook_from_path`.

    Returns:
        O :class:`openpyxl.Workbook` carregado.

    Raises:
        InvalidWorkbookError: Se a extensão for inválida ou o conteúdo não for
            um Excel legível.
    """
    _validate_extension(filename)
    if keep_vba is None:
        keep_vba = _should_keep_vba(filename)

    buffer: IO[bytes] = io.BytesIO(data) if isinstance(data, (bytes, bytearray)) else data

    logger.info("Carregando workbook de bytes '%s' (keep_vba=%s)", filename, keep_vba)
    try:
        return load_workbook(filename=buffer, data_only=data_only, keep_vba=keep_vba)
    except InvalidFileException as exc:
        raise InvalidWorkbookError(f"Arquivo Excel inválido: {filename}") from exc
    except Exception as exc:  # noqa: BLE001 - normaliza qualquer falha de leitura
        raise InvalidWorkbookError(
            f"Não foi possível abrir '{filename}': {exc}"
        ) from exc


def save_workbook_to_bytes(workbook: Workbook) -> io.BytesIO:
    """Serializa o workbook em memória, sem tocar em arquivos do disco.

    Args:
        workbook: Workbook a salvar.

    Returns:
        Um :class:`io.BytesIO` posicionado no início, pronto para download.

    Raises:
        InvalidWorkbookError: Se a serialização falhar.
    """
    buffer = io.BytesIO()
    try:
        workbook.save(buffer)
    except Exception as exc:  # noqa: BLE001
        raise InvalidWorkbookError(f"Falha ao gerar o arquivo final: {exc}") from exc
    buffer.seek(0)
    return buffer
