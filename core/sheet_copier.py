"""Núcleo técnico: cópia integral de abas entre workbooks.

⚠️ Ponto crítico do ``openpyxl``: objetos de estilo (``Font``, ``Fill``,
``Border``…) **não podem ser reaproveitados entre workbooks diferentes**, pois
internamente são indexados na tabela de estilos de cada workbook. Por isso todo
estilo é clonado com :func:`copy.copy`, o que faz o ``openpyxl`` registrá-lo
corretamente na tabela do workbook de destino.

A cópia é *resiliente*: o conteúdo das células (valores, fórmulas e estilos) é
tratado como crítico — uma falha aqui aborta a aba com
:class:`~core.exceptions.CopyOperationError`. Já os aspectos estruturais
(dimensões, mescladas, validações, gráficos…) são *best-effort*: uma falha em um
deles é registrada em ``warnings`` e a cópia segue, sem perder os dados.
"""

from __future__ import annotations

import copy
import logging
from dataclasses import dataclass, field
from typing import Callable

from openpyxl.comments import Comment
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.workbook import Workbook
from openpyxl.worksheet.datavalidation import DataValidationList
from openpyxl.worksheet.worksheet import Worksheet

from config.defaults import COPY_MODE_OVERLAY, COPY_MODE_REPLACE
from core.exceptions import CopyOperationError, SheetNotFoundError

logger = logging.getLogger(__name__)

ProgressCallback = Callable[[int, int, str], None]


# --------------------------------------------------------------------------- #
# Resultados
# --------------------------------------------------------------------------- #
@dataclass
class SheetCopyResult:
    """Resultado da cópia de uma aba.

    Attributes:
        sheet_name: Nome da aba copiada.
        created: ``True`` se a aba não existia na destino e foi criada.
        mode: Modo de cópia aplicado (``replace`` ou ``overlay``).
        cells_copied: Número de células com conteúdo/estilo copiadas.
        merged_ranges: Intervalos mesclados aplicados.
        data_validations: Validações de dados copiadas.
        conditional_formats: Regras de formatação condicional copiadas.
        charts: Gráficos copiados.
        images: Imagens copiadas.
        warnings: Mensagens de aspectos que falharam (best-effort).
    """

    sheet_name: str
    created: bool
    mode: str
    cells_copied: int = 0
    merged_ranges: int = 0
    data_validations: int = 0
    conditional_formats: int = 0
    charts: int = 0
    images: int = 0
    warnings: list[str] = field(default_factory=list)


@dataclass
class CopyReport:
    """Relatório consolidado de uma operação de cópia de várias abas."""

    results: list[SheetCopyResult] = field(default_factory=list)
    added_named_ranges: list[str] = field(default_factory=list)

    @property
    def sheets_copied(self) -> int:
        """Total de abas processadas."""
        return len(self.results)

    @property
    def sheets_created(self) -> int:
        """Quantas abas foram criadas (não existiam na destino)."""
        return sum(1 for r in self.results if r.created)

    @property
    def total_warnings(self) -> int:
        """Soma de avisos em todas as abas."""
        return sum(len(r.warnings) for r in self.results)


# --------------------------------------------------------------------------- #
# Helpers internos
# --------------------------------------------------------------------------- #
def _unique_temp_name(wb: Workbook) -> str:
    """Gera um nome de aba temporário inexistente no workbook (≤ 31 chars)."""
    base = "__tmp_copy__"
    candidate = base
    i = 0
    while candidate in wb.sheetnames:
        i += 1
        candidate = f"{base}{i}"
    return candidate[:31]


def _recreate_sheet(target_wb: Workbook, name: str) -> Worksheet:
    """Recria a aba ``name`` na mesma posição, com conteúdo zerado.

    Renomeia a antiga para um nome temporário antes de remover, garantindo que
    o workbook nunca fique sem nenhuma aba (caso de borda do ``openpyxl``).
    """
    idx = target_wb.sheetnames.index(name)
    old_ws = target_wb[name]
    tmp_name = _unique_temp_name(target_wb)
    old_ws.title = tmp_name
    new_ws = target_wb.create_sheet(title=name, index=idx)
    del target_wb[tmp_name]
    return new_ws


def _clone_comment(comment: Comment) -> Comment:
    """Clona um comentário (um ``Comment`` não pode pertencer a duas células)."""
    new = Comment(text=comment.text or "", author=comment.author or "")
    new.width = comment.width
    new.height = comment.height
    return new


def _copy_one_cell(source_cell, target_ws: Worksheet) -> bool:
    """Copia valor, estilo, comentário e hyperlink de uma célula.

    Returns:
        ``True`` se algo foi escrito; ``False`` para células totalmente vazias
        (otimização: evita inflar o arquivo destino com células default).
    """
    has_content = (
        source_cell.value is not None
        or source_cell.has_style
        or source_cell.comment is not None
        or source_cell.hyperlink is not None
    )
    if not has_content:
        return False

    new_cell = target_ws.cell(row=source_cell.row, column=source_cell.column)
    new_cell.value = source_cell.value

    if source_cell.has_style:
        new_cell.font = copy.copy(source_cell.font)
        new_cell.fill = copy.copy(source_cell.fill)
        new_cell.border = copy.copy(source_cell.border)
        new_cell.alignment = copy.copy(source_cell.alignment)
        new_cell.number_format = source_cell.number_format
        new_cell.protection = copy.copy(source_cell.protection)

    if source_cell.comment is not None:
        new_cell.comment = _clone_comment(source_cell.comment)

    if source_cell.hyperlink is not None:
        new_cell.hyperlink = copy.copy(source_cell.hyperlink)

    return True


def _copy_cells(source_ws: Worksheet, target_ws: Worksheet) -> int:
    """Copia todas as células não vazias do intervalo usado. Retorna a contagem."""
    count = 0
    for row in source_ws.iter_rows():
        for cell in row:
            if _copy_one_cell(cell, target_ws):
                count += 1
    return count


def _unmerge_all(target_ws: Worksheet) -> None:
    """Remove todas as mescladas da aba destino (necessário no modo overlay)."""
    for rng in list(target_ws.merged_cells.ranges):
        target_ws.unmerge_cells(str(rng))


def _apply_merges(source_ws: Worksheet, target_ws: Worksheet) -> int:
    """Aplica na destino as mescladas da origem. Retorna a contagem."""
    count = 0
    for rng in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(rng))
        count += 1
    return count


def _copy_dimensions(source_ws: Worksheet, target_ws: Worksheet) -> None:
    """Copia larguras de coluna e alturas de linha (incluindo agrupamentos)."""
    for key, dim in source_ws.column_dimensions.items():
        tgt = target_ws.column_dimensions[key]
        if dim.width is not None:
            tgt.width = dim.width  # define `width` já liga `customWidth` (read-only)
        tgt.hidden = dim.hidden
        tgt.outline_level = dim.outline_level
        tgt.bestFit = dim.bestFit
        tgt.collapsed = dim.collapsed
        tgt.min = dim.min
        tgt.max = dim.max

    for key, dim in source_ws.row_dimensions.items():
        tgt = target_ws.row_dimensions[key]
        if dim.height is not None:
            tgt.height = dim.height
        tgt.hidden = dim.hidden
        tgt.outline_level = dim.outline_level
        tgt.collapsed = dim.collapsed


def _copy_data_validations(source_ws: Worksheet, target_ws: Worksheet) -> int:
    """Substitui as validações de dados da destino pelas da origem."""
    target_ws.data_validations = DataValidationList()
    count = 0
    for dv in source_ws.data_validations.dataValidation:
        target_ws.add_data_validation(copy.copy(dv))
        count += 1
    return count


def _copy_conditional_formatting(source_ws: Worksheet, target_ws: Worksheet) -> int:
    """Copia formatação condicional, resolvendo o ``dxf`` entre workbooks.

    Regras carregadas de arquivo guardam apenas o ``dxfId`` (índice na tabela do
    workbook de origem). Resolvemos o ``DifferentialStyle`` real e zeramos o
    ``dxfId`` para que o workbook de destino atribua um índice novo e correto —
    do contrário as cores da formatação condicional sairiam trocadas.
    """
    target_ws.conditional_formatting = ConditionalFormattingList()
    src_dxfs = getattr(source_ws.parent, "_differential_styles", None)
    cf_rules = getattr(source_ws.conditional_formatting, "_cf_rules", {})

    count = 0
    for cf, rules in cf_rules.items():
        sqref = str(cf.sqref)
        for rule in rules:
            new_rule = copy.copy(rule)
            if new_rule.dxf is None and new_rule.dxfId is not None and src_dxfs is not None:
                try:
                    new_rule.dxf = copy.copy(src_dxfs[new_rule.dxfId])
                except (IndexError, TypeError, KeyError):
                    pass
            new_rule.dxfId = None
            target_ws.conditional_formatting.add(sqref, new_rule)
            count += 1
    return count


def _copy_layout(source_ws: Worksheet, target_ws: Worksheet, warnings: list[str]) -> None:
    """Copia congelamento, filtro, cor da aba, visão e configurações de página."""
    target_ws.freeze_panes = source_ws.freeze_panes
    if source_ws.auto_filter.ref:
        target_ws.auto_filter.ref = source_ws.auto_filter.ref
    target_ws.sheet_properties.tabColor = source_ws.sheet_properties.tabColor
    target_ws.sheet_state = source_ws.sheet_state

    try:
        s_view, t_view = source_ws.sheet_view, target_ws.sheet_view
        t_view.showGridLines = s_view.showGridLines
        t_view.zoomScale = s_view.zoomScale
        t_view.zoomScaleNormal = s_view.zoomScaleNormal
        t_view.rightToLeft = s_view.rightToLeft
        t_view.showRowColHeaders = s_view.showRowColHeaders
    except Exception as exc:  # noqa: BLE001
        warnings.append(f"visão da aba: {exc}")

    try:
        target_ws.page_setup = copy.copy(source_ws.page_setup)
        target_ws.print_options = copy.copy(source_ws.print_options)
        target_ws.page_margins = copy.copy(source_ws.page_margins)
    except Exception as exc:  # noqa: BLE001
        warnings.append(f"configuração de página: {exc}")


def _copy_charts(source_ws: Worksheet, target_ws: Worksheet, warnings: list[str]) -> int:
    """Copia gráficos via ``deepcopy`` (preserva séries e âncora). Resiliente."""
    target_ws._charts = []  # zera para não duplicar no modo overlay
    count = 0
    for chart in getattr(source_ws, "_charts", []):
        try:
            target_ws.add_chart(copy.deepcopy(chart))
            count += 1
        except Exception as exc:  # noqa: BLE001
            warnings.append(f"gráfico: {exc}")
    return count


def _copy_images(source_ws: Worksheet, target_ws: Worksheet, warnings: list[str]) -> int:
    """Copia imagens via ``deepcopy`` (preserva âncora). Resiliente."""
    target_ws._images = []  # zera para não duplicar no modo overlay
    count = 0
    for img in getattr(source_ws, "_images", []):
        try:
            target_ws.add_image(copy.deepcopy(img))
            count += 1
        except Exception as exc:  # noqa: BLE001
            warnings.append(f"imagem: {exc}")
    return count


# --------------------------------------------------------------------------- #
# API pública
# --------------------------------------------------------------------------- #
def copy_sheet_full(
    source_ws: Worksheet,
    target_wb: Workbook,
    mode: str = COPY_MODE_REPLACE,
) -> SheetCopyResult:
    """Copia uma aba inteira da origem para o workbook de destino.

    Args:
        source_ws: Worksheet de origem (do modelo).
        target_wb: Workbook de destino (será modificado in-place).
        mode: ``replace`` (recria a aba do zero — padrão) ou ``overlay``
            (mantém a aba e sobrescreve apenas as células do modelo).

    Returns:
        Um :class:`SheetCopyResult` com contagens e avisos.

    Raises:
        CopyOperationError: Se a cópia das células (conteúdo crítico) falhar.
    """
    name = source_ws.title
    warnings: list[str] = []
    existed = name in target_wb.sheetnames
    created = not existed
    logger.info("Copiando aba '%s' (modo=%s, criada=%s)", name, mode, created)

    try:
        if existed and mode == COPY_MODE_REPLACE:
            target_ws = _recreate_sheet(target_wb, name)
        elif existed:  # COPY_MODE_OVERLAY
            target_ws = target_wb[name]
        else:
            target_ws = target_wb.create_sheet(name)

        # Desmesclar a destino ANTES de escrever células (no overlay, células
        # dentro de uma mescla são read-only e quebrariam a escrita).
        try:
            _unmerge_all(target_ws)
        except Exception as exc:  # noqa: BLE001
            warnings.append(f"desmesclar destino: {exc}")

        # CRÍTICO: conteúdo das células.
        cells_copied = _copy_cells(source_ws, target_ws)

        # Best-effort: aspectos estruturais.
        try:
            _copy_dimensions(source_ws, target_ws)
        except Exception as exc:  # noqa: BLE001
            warnings.append(f"dimensões: {exc}")

        try:
            merged = _apply_merges(source_ws, target_ws)
        except Exception as exc:  # noqa: BLE001
            warnings.append(f"mescladas: {exc}")
            merged = 0

        try:
            dvs = _copy_data_validations(source_ws, target_ws)
        except Exception as exc:  # noqa: BLE001
            warnings.append(f"validações de dados: {exc}")
            dvs = 0

        try:
            cfs = _copy_conditional_formatting(source_ws, target_ws)
        except Exception as exc:  # noqa: BLE001
            warnings.append(f"formatação condicional: {exc}")
            cfs = 0

        _copy_layout(source_ws, target_ws, warnings)
        charts = _copy_charts(source_ws, target_ws, warnings)
        images = _copy_images(source_ws, target_ws, warnings)

    except CopyOperationError:
        raise
    except Exception as exc:  # noqa: BLE001
        raise CopyOperationError(f"Falha ao copiar a aba '{name}': {exc}") from exc

    return SheetCopyResult(
        sheet_name=name,
        created=created,
        mode=mode,
        cells_copied=cells_copied,
        merged_ranges=merged,
        data_validations=dvs,
        conditional_formats=cfs,
        charts=charts,
        images=images,
        warnings=warnings,
    )


def copy_defined_names(source_wb: Workbook, target_wb: Workbook) -> list[str]:
    """Copia named ranges da origem para a destino, sem sobrescrever existentes.

    Args:
        source_wb: Workbook de origem.
        target_wb: Workbook de destino (modificado in-place).

    Returns:
        Lista dos nomes adicionados (os já existentes na destino são mantidos).
    """
    added: list[str] = []
    for name in list(source_wb.defined_names):
        if name not in target_wb.defined_names:
            target_wb.defined_names[name] = copy.copy(source_wb.defined_names[name])
            added.append(name)
    if added:
        logger.info("Named ranges adicionados: %s", ", ".join(added))
    return added


def copy_sheets(
    source_wb: Workbook,
    target_wb: Workbook,
    sheet_names: list[str],
    *,
    mode: str = COPY_MODE_REPLACE,
    copy_named_ranges: bool = True,
    progress_callback: ProgressCallback | None = None,
) -> CopyReport:
    """Copia várias abas da origem para a destino e devolve um relatório.

    Args:
        source_wb: Workbook de origem (modelo).
        target_wb: Workbook de destino (modificado in-place).
        sheet_names: Nomes das abas do modelo a copiar.
        mode: ``replace`` ou ``overlay`` (veja :func:`copy_sheet_full`).
        copy_named_ranges: Se ``True``, copia named ranges ausentes na destino.
        progress_callback: Chamado após cada aba como ``(atual, total, nome)``,
            útil para barras de progresso na UI.

    Returns:
        Um :class:`CopyReport` consolidado.

    Raises:
        SheetNotFoundError: Se alguma aba pedida não existir no modelo.
        CopyOperationError: Se a cópia de conteúdo de alguma aba falhar.
    """
    if mode not in (COPY_MODE_REPLACE, COPY_MODE_OVERLAY):
        raise ValueError(f"Modo de cópia inválido: {mode!r}")

    results: list[SheetCopyResult] = []
    total = len(sheet_names)
    for i, name in enumerate(sheet_names, start=1):
        if name not in source_wb.sheetnames:
            raise SheetNotFoundError(f"A aba '{name}' não existe no modelo.")
        results.append(copy_sheet_full(source_wb[name], target_wb, mode=mode))
        if progress_callback is not None:
            progress_callback(i, total, name)

    added = copy_defined_names(source_wb, target_wb) if copy_named_ranges else []
    return CopyReport(results=results, added_named_ranges=added)
