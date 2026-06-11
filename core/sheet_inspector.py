"""Extração de metadados das abas de um workbook.

Fornece à interface os números exibidos ao lado de cada aba (fórmulas,
validações, células mescladas, gráficos…), sem alterar o workbook.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class SheetMetadata:
    """Resumo quantitativo de uma aba.

    Attributes:
        name: Nome da aba.
        formula_count: Número de células contendo fórmulas.
        data_validation_count: Número de validações de dados.
        merged_cells_count: Número de intervalos mesclados.
        chart_count: Número de gráficos.
        image_count: Número de imagens.
        conditional_format_count: Número de regras de formatação condicional.
        max_row: Última linha usada.
        max_column: Última coluna usada.
        hidden: Se a aba está oculta.
    """

    name: str
    formula_count: int = 0
    data_validation_count: int = 0
    merged_cells_count: int = 0
    chart_count: int = 0
    image_count: int = 0
    conditional_format_count: int = 0
    max_row: int = 0
    max_column: int = 0
    hidden: bool = False

    def summary_line(self) -> str:
        """Retorna uma descrição curta (ex.: '242 fórmulas, 5 mescladas, 1 gráfico').

        Inclui apenas os atributos com valor maior que zero. Retorna
        ``'(vazia)'`` quando não há nada relevante a destacar.
        """
        parts: list[str] = []
        if self.formula_count:
            parts.append(f"{self.formula_count} fórmulas")
        if self.data_validation_count:
            parts.append(f"{self.data_validation_count} validações")
        if self.merged_cells_count:
            parts.append(f"{self.merged_cells_count} mescladas")
        if self.conditional_format_count:
            parts.append(f"{self.conditional_format_count} format. condic.")
        if self.chart_count:
            sufixo = "gráfico" if self.chart_count == 1 else "gráficos"
            parts.append(f"{self.chart_count} {sufixo}")
        if self.image_count:
            sufixo = "imagem" if self.image_count == 1 else "imagens"
            parts.append(f"{self.image_count} {sufixo}")
        return ", ".join(parts) if parts else "(sem fórmulas)"


def _count_formulas(ws: Worksheet) -> int:
    """Conta células com fórmula (``data_type == 'f'``) no intervalo usado."""
    total = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == "f":
                total += 1
    return total


def _count_conditional_formats(ws: Worksheet) -> int:
    """Conta regras de formatação condicional de forma defensiva."""
    cf = ws.conditional_formatting
    rules = getattr(cf, "_cf_rules", None)
    if rules is None:
        return 0
    return sum(len(rule_list) for rule_list in rules.values())


def inspect_sheet(ws: Worksheet) -> SheetMetadata:
    """Inspeciona uma aba e devolve seus metadados.

    Args:
        ws: A worksheet a inspecionar.

    Returns:
        Um :class:`SheetMetadata` preenchido.
    """
    return SheetMetadata(
        name=ws.title,
        formula_count=_count_formulas(ws),
        data_validation_count=len(ws.data_validations.dataValidation),
        merged_cells_count=len(ws.merged_cells.ranges),
        chart_count=len(getattr(ws, "_charts", [])),
        image_count=len(getattr(ws, "_images", [])),
        conditional_format_count=_count_conditional_formats(ws),
        max_row=ws.max_row,
        max_column=ws.max_column,
        hidden=ws.sheet_state != "visible",
    )


def inspect_workbook(wb: Workbook) -> list[SheetMetadata]:
    """Inspeciona todas as abas do workbook, na ordem em que aparecem.

    Args:
        wb: O workbook a inspecionar.

    Returns:
        Lista de :class:`SheetMetadata`, uma por aba.
    """
    metadata = [inspect_sheet(wb[name]) for name in wb.sheetnames]
    logger.info("Inspecionadas %d abas", len(metadata))
    return metadata


@dataclass(frozen=True)
class WorkbookSummary:
    """Resumo de alto nível de um workbook (para o cabeçalho da UI)."""

    sheet_count: int
    sheet_names: list[str] = field(default_factory=list)


def workbook_summary(wb: Workbook) -> WorkbookSummary:
    """Retorna contagem e nomes das abas do workbook."""
    return WorkbookSummary(sheet_count=len(wb.sheetnames), sheet_names=list(wb.sheetnames))
