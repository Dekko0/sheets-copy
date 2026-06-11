"""Detecção de dependências cruzadas entre abas.

Antes de copiar, varremos as fórmulas das abas selecionadas em busca de
referências a outras abas (ex.: ``=FATURAS!M1``). Uma dependência é considerada
**atendida** se a aba referenciada já existe na destino *ou* também está sendo
copiada nesta operação; caso contrário é **quebrada**.

A varredura ignora referências dentro de literais de texto (ex.:
``="Olá!"``) e referências à própria aba.
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field

from openpyxl.workbook import Workbook

logger = logging.getLogger(__name__)

# Remove literais de texto ("...") para não confundir um '!' dentro de uma
# string com uma referência de aba. O Excel escapa aspas duplas como "".
_STRING_LITERAL_RE = re.compile(r'"(?:[^"]|"")*"')

# Captura "Aba!" — quotada ('Nome Com Espaço'!) ou não (Aba_1!).
# Nomes não-quotados aceitam letras acentuadas (intervalo latino) usadas em PT-BR.
_QUOTED = r"'((?:[^']|'')+)'"
_UNQUOTED = r"([A-Za-z0-9_.À-ɏ]+)"
_SHEET_REF_RE = re.compile(rf"(?:{_QUOTED}|{_UNQUOTED})!")


def _normalize(name: str) -> str:
    """Normaliza um nome de aba para comparação (Excel é case-insensitive)."""
    return name.strip().casefold()


def extract_sheet_references(formula: str) -> set[str]:
    """Extrai os nomes de abas referenciados em uma fórmula.

    Args:
        formula: Texto da fórmula (com ou sem o ``=`` inicial).

    Returns:
        Conjunto de nomes de abas referenciados (sem aspas, com ``''`` já
        convertido para ``'``). Vazio se não houver referências.

    Examples:
        >>> sorted(extract_sheet_references("=FATURAS!A1 + Dados!B2"))
        ['Dados', 'FATURAS']
        >>> extract_sheet_references('="texto com ! dentro"')
        set()
    """
    if not formula:
        return set()

    clean = _STRING_LITERAL_RE.sub("", formula)
    refs: set[str] = set()
    for quoted, unquoted in _SHEET_REF_RE.findall(clean):
        name = quoted.replace("''", "'") if quoted else unquoted
        if name:
            refs.add(name)
    return refs


def _cell_formula(cell) -> str | None:
    """Devolve o texto da fórmula de uma célula, ou ``None`` se não for fórmula."""
    if cell.data_type != "f":
        return None
    value = cell.value
    if isinstance(value, str):
        return value
    # Fórmulas de matriz (ArrayFormula) guardam o texto em ``.text``.
    return getattr(value, "text", None)


@dataclass(frozen=True)
class SheetDependency:
    """Uma referência de ``source_sheet`` para ``referenced_sheet``.

    Attributes:
        source_sheet: Aba (sendo copiada) que contém a fórmula.
        referenced_sheet: Aba referenciada pela fórmula.
        example_cell: Coordenada de um exemplo (ex.: ``"B1"``).
        example_formula: Texto (possivelmente truncado) da fórmula de exemplo.
    """

    source_sheet: str
    referenced_sheet: str
    example_cell: str
    example_formula: str


@dataclass
class DependencyReport:
    """Resultado da análise de dependências."""

    satisfied: list[SheetDependency] = field(default_factory=list)
    broken: list[SheetDependency] = field(default_factory=list)

    @property
    def has_broken(self) -> bool:
        """``True`` se houver ao menos uma dependência quebrada."""
        return bool(self.broken)

    @property
    def broken_sheets(self) -> list[str]:
        """Abas faltantes na destino, em ordem alfabética e sem repetição."""
        return sorted({dep.referenced_sheet for dep in self.broken})


def _truncate(text: str, limit: int = 120) -> str:
    """Trunca uma fórmula longa para exibição."""
    text = text.strip()
    return text if len(text) <= limit else text[: limit - 1] + "…"


def analyze_dependencies(
    source_wb: Workbook,
    target_wb: Workbook,
    sheet_names: list[str],
    *,
    ignore_self: bool = True,
) -> DependencyReport:
    """Analisa dependências das abas selecionadas contra a planilha destino.

    Uma dependência é atendida quando a aba referenciada já existe na destino ou
    está entre as abas que serão copiadas (pois serão criadas na destino).

    Args:
        source_wb: Workbook de origem (modelo).
        target_wb: Workbook de destino.
        sheet_names: Abas do modelo selecionadas para cópia.
        ignore_self: Se ``True`` (padrão), ignora referências de uma aba a si
            mesma.

    Returns:
        Um :class:`DependencyReport` com listas de dependências atendidas e
        quebradas (uma entrada por par aba-origem → aba-referenciada).
    """
    # Abas que existirão na destino APÓS a cópia.
    available = {_normalize(s) for s in target_wb.sheetnames}
    available |= {_normalize(s) for s in sheet_names}

    seen: set[tuple[str, str]] = set()
    report = DependencyReport()

    for name in sheet_names:
        if name not in source_wb.sheetnames:
            continue
        ws = source_wb[name]
        for row in ws.iter_rows():
            for cell in row:
                formula = _cell_formula(cell)
                if not formula:
                    continue
                for ref in extract_sheet_references(formula):
                    if ignore_self and _normalize(ref) == _normalize(name):
                        continue
                    pair = (_normalize(name), _normalize(ref))
                    if pair in seen:
                        continue
                    seen.add(pair)

                    dependency = SheetDependency(
                        source_sheet=name,
                        referenced_sheet=ref,
                        example_cell=cell.coordinate,
                        example_formula=_truncate(formula),
                    )
                    if _normalize(ref) in available:
                        report.satisfied.append(dependency)
                    else:
                        report.broken.append(dependency)

    logger.info(
        "Dependências: %d atendidas, %d quebradas",
        len(report.satisfied),
        len(report.broken),
    )
    return report
